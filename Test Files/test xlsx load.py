from openpyxl import load_workbook

wb = load_workbook(filename='test.xlsx')

ws2 = wb.create_sheet(title="Pi")

ws2['F5'] = 3.14

wb.save(filename = 'success.xlsx')

print('Lets go look!')



