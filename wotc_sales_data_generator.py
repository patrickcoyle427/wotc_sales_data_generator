#!/usr/bin/env python3

'''
wotc_sales_data_generator.py

Takes a CSV from lightspeed retail POS, pulls out the sales information that Wizards
Of the Coast needs, then writes them to an xlsx file that can be sent for reporting.

-Written by Patrick Coyle

USAGE:

- REQUIRED DATA: Sales by line exported to .csv files in Lightspeed Retail for the month
  you wish to parse. You will need 2 .csv files to cover all the data needed by wotc, one for
  the mtg tag and one for dnd and exclude minis (they aren't made by wotc so they aren't
  included)

  These required files can be found the following way
  1. Go to Reports
  2. Go to Lines
  3. Change Shop to Blue Bell
  4. Change the dates to all those in the month you wish to scan
  5. Use the tags to filter results (mtg for Magic: The gathering, dnd for Dungeons and Dragons)
  6. Search each tag one at a time and export the results, the export button is on the
      right side of the Lines screen
  
- On first run of this script, two directories will be created in the same directory as this
  file: to_parse and parsed_data

- Place Lightspeed Retail POS sales data into the to_parse folder, these files are .csv files
  that are created in sales by line in the reporting section of Lightspeed Retail POS

- WOTC wants 1 month of all Magic: The Gathering and Dungeons and Dragons Sales. To create
  the data that will be used for this script, In Lightspeed POS, go to Reports, then Sales
  Lines. Set the date range to be the month you want sales from. For Magic: The Gathering
  search the "MTG" tag, and exclude "accessory". For Dungeons and Dragons, search the "dnd"
  tag and exclude "mini".

'''

#TO DO

# More error handing when there is user input
# More error handing when files are read or created
# Add the ability to type delte for a UPC and have the parser skip over it
#   When the report is written
# Maybe change the output format to .csv instead of .xlsx to cut down
#   on imported moduals

import csv, os, os.path, shutil

from datetime import datetime
from openpyxl import load_workbook

# csv - reads the data from Lightspeed Retail POS
# os - used for creating folders and checking folder contents
# os.path - Used for checking if the 'to_parse' and 'parsed_files'
#           folders exist
# shutil - used to create a copy of the templat file to append.
# from datetime importdate - used to get the month and year to create the file name of the output file
# from openpyxl import load_workbook - used to read, modify and save the results in the .xlsx format

def start_parse():

    # Runs script through each step of the parsing process

    dir_exist = dir_check()

    if dir_exist:

        findings = find_csv_names()

        pulled_data = pull_data(findings)

        generated_report = generate_report(pulled_data)

        if generated_report:

            print('Report successfully generated')

def dir_check():

    # Will create necessary directories on first launch of script,
    # then will pass each other time. Gives a message to let the user know
    # what to do with the created directories

    if os.path.isdir('to_parse') and os.path.isdir('parsed_files'):

        return True

    else:

        create_these = ('to_parse', 'parsed_files')

        for i in create_these:

            if not os.path.isdir(i):

                os.makedirs(i)

        print('Folders for parsing created. Please place all files that need to be parsed',
              'into the "to_parse" folder, then run this script again.')

        return False

def find_csv_names():

    # Finds the names of the files to be parsed.
    
    return [file for file in os.listdir('to_parse') if file.endswith('.csv')]

def pull_data(names):

    # Takes the data from each CSV and loads it into memory to merged and written
    # to a single .xlsx file.

    # names - the file names of each file that needs to be parsed

    wotc_id = '5676'
    # Alternate Universes's wotc ID
    # WOTC ID number is the first column in the spreadsheet that will be written

    new_upcs = {}
    new_upcs_added = False
    # When new_upcs is set to true, the wotc_master_upc.csv file gets created if it doesn't exist
    # or appended if it already does

    data = []

    master_upc = load_master_upc()
    # load_master_upc() loads all previously recorded UPCs

    user_upc = ''

    for file in names:

        with open(f'to_parse/{file}', newline='') as csvfile:

            reader = csv.reader(csvfile, delimiter=',')

            next(reader, None)
            # Skips the header line

            for row in reader:

                transaction_date = row[1]
                transaction_id = row[0]
                prod_name = row[2]
                qty = row[3]
                retail_cost = row[4]
                subtotal = row[5]

                upc = ''

                if prod_name not in new_upcs and prod_name not in master_upc:
                
                    # ask user for upc

                    while user_upc == '':

                        # prevents the user from accidently inputing a blank line

                        try:

                            user_upc = input(f'What is the UPC of {prod_name}?\n>  ')

                            # To DO: Let the user Type DELETE to remove the item record

                        except KeyboardInterrupt:

                            print('\nPlease Enter a upc')

                            continue

                

                    new_upcs[prod_name] = user_upc

                    new_upcs_added = True

                    upc = user_upc

                    user_upc = ''
                    # resets the user UPC
                
                else:

                    if prod_name in new_upcs:

                        upc = new_upcs[prod_name]

                    else:

                        upc = master_upc[prod_name]

                if upc.lower() != 'delete':

                    data.append((wotc_id, transaction_date, transaction_id, upc, prod_name, qty, retail_cost, subtotal))

    if new_upcs_added:

        status = export_new_upcs(new_upcs)

        if status:

            print('wotc_master_upcs.csv successfully updated')

    move_parsed_files(names)

    return data

def move_parsed_files(names):

    # Moves files that were parsed for the report to a new folder so the user knows they've been scanned

    # names - a list of file names that have been parsed

    for file in names:

        os.replace(f'to_parse/{file}',  f'parsed_files/{file}')

def load_master_upc():

    # Loads the master UPC file data, used for checking previously recorded UPCs
    # returns master_upcs, which is a dictionary with all the previously recorded UPCs
    
    master_upc = {}
    
    if os.path.exists('wotc_master_upc.csv'):

        with open('wotc_master_upc.csv', newline='') as known_upcs:

           upc_reader = csv.reader(known_upcs)

           next(upc_reader, None)
           # Skips the header line

           for row in upc_reader:

               loaded_prod = row[0]
               loaded_upc = row[1]

               master_upc[loaded_prod] = loaded_upc

    return master_upc
    

def generate_report(data):

    # Loads the template file, writes data to a copy of it and then saves it in the format that WOTC requires

    # data - table that contains tuples with all the information to go onto the report in the correct order
    #        that was created in the pull_data function.

    if os.path.exists('wotc_report_template.xlsx'):

        column_letters = ('A', 'B', 'C', 'E', 'I', 'J', 'K', 'L')
        # Letters of the columns used

        today = datetime.today()
        # used for building a file name

        year = str(today.year)[2:]
        # Only needs the last 2 digits of the year

        month = str(today.month - 1)
        
        if int(month) == 0:

            month = '12'
            # Month 0 doesnt exist, it means december aka 12
            
            year = str(int(year) - 1)
            # Changes the year too

            # TO DO: make some corner case years work

        elif len(month) == 1:

            month = f'0{month}'
            # wotc file formatting requires the month to have a 0 if necessary

        new_data_name = f'NEW_5676_AlternateUniversesEastNorriton_POSdata_{month}{year}.xlsx'

        shutil.copy('wotc_report_template.xlsx', new_data_name)        

    else:

        print('wotc_report_template.xlsx is missing. Please locate this file and place it in the same folder')
        print('as "wotc_sales_data_generator.py"!')

        return False


    wb = load_workbook(filename=new_data_name)

    ws = wb.active

    start_row = 5

    row_position = 0
    
    for row_num, row_data in enumerate(data):

        for column in column_letters:

            ws[column + str(row_num + start_row)] = row_data[row_position]

            row_position += 1

        row_position = 0

    wb.save(new_data_name)

    return True
    
def export_new_upcs(upc_dict):

    # Exports the UPC dict to be used on subsequent runs of the script.
    # Will only export if new changes are made.

    # upc_dict - Dictionary created in pull data with the UPCs of products scanned

    # TODO: Add error handling

    csv_columns = ('Item Name', 'UPC')

    write_header = False

    file_name = 'wotc_master_upc.csv'

    if os.path.exists(file_name) == False:

        write_header = True     

    with open(file_name, 'a', newline='') as master_upc:

        writer = csv.DictWriter(master_upc, fieldnames=csv_columns)

        if write_header == True:

            writer.writeheader()

        for key in upc_dict.keys():

            writer.writerow({csv_columns[0]: key, csv_columns[1]: upc_dict[key]})

    return True
        

if __name__ == '__main__':

    start_parse()
